### Variables
file_in = input("Enter Workbook File Location")
omit_header = input("Omit First Row? [yes = 1, no = 0]")
omit_header= int(omit_header)


### Imports
import openpyxl as xl
from pathlib import Path

### Open Existing Workbook + Create New Workbook
old_wb = xl.load_workbook(file_in)
new_wb = xl.Workbook()
new_ws = new_wb.active

old_sheets = old_wb.sheetnames


print(old_sheets)

row_count = 1
for old_sheet in old_sheets:
    for row in range (1 + omit_header, old_wb[old_sheet].max_row + 1):
        for col in range(1, old_wb[old_sheet].max_column + 1):
            new_ws.cell(row = row_count, column = col).value = old_wb[old_sheet].cell(row = row, column = col).value
        row_count +=1




new_wb.save(str(Path(file_in).parent/Path(file_in).stem) + '_edited' + str(Path(file_in).suffix))
